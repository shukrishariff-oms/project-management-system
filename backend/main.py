from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
import bcrypt
from jose import JWTError, jwt
from datetime import datetime, timedelta, date
from typing import Optional
import openpyxl
from io import BytesIO
import tempfile
import os

import models, schemas, database

# Create tables
models.Base.metadata.create_all(bind=database.engine)

# Run manual migration for schema updates
# import manual_migration
# manual_migration.migrate(database.engine)

app = FastAPI()

# CORS Setup
origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security Config
SECRET_KEY = "supersecretkey" # Change this in production
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Helper Functions
def verify_password(plain_password, hashed_password):
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def get_password_hash(password):
    # Hash password using bcrypt
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def recalculate_project_progress(project_id: int, db: Session):
    """Calculate project progress based on completed tasks"""
    tasks = db.query(models.ProjectTask).filter(models.ProjectTask.project_id == project_id).all()
    if not tasks:
        progress = 0
    else:
        completed = [t for t in tasks if t.status == "Completed"]
        progress = int((len(completed) / len(tasks)) * 100)
    
    db.query(models.Project).filter(models.Project.id == project_id).update({"progress_percentage": progress})
    db.commit()

# Routes
@app.post("/users", response_model=schemas.User)
def create_user(user: schemas.UserCreate, db: Session = Depends(database.get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed_password = get_password_hash(user.password)
    db_user = models.User(
        email=user.email, 
        password_hash=hashed_password,
        full_name=user.full_name,
        role=user.role
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.post("/login", response_model=schemas.Token)
def login(user: schemas.UserLogin, db: Session = Depends(database.get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if not db_user or not verify_password(user.password, db_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": db_user.email}, expires_delta=access_token_expires
    )
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "role": db_user.role,
        "full_name": db_user.full_name,
        "email": db_user.email
    }

@app.get("/")
def read_root():
    return {"message": "Welcome to Project Management System API"}

# Project Routes
@app.post("/api/projects", response_model=schemas.Project)
def create_project(project: schemas.ProjectCreate, db: Session = Depends(database.get_db)):
    """Create a new project with default health metrics"""
    # Create the project
    db_project = models.Project(
        name=project.name,
        project_code=project.project_code,
        project_manager=project.project_manager,
        assigned_to_email=project.assigned_to_email,
        start_date=project.start_date,
        end_date=project.end_date,
        planned_cost=project.planned_cost,
        actual_cost=project.actual_cost,
        description=project.description,
        objective=project.objective,
        status=project.status,
        priority=project.priority,
        risk_level=project.risk_level,
        department=project.department,
        is_archived=project.is_archived,
        tags=project.tags,
        progress_percentage=project.progress_percentage
    )
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    
    # Initialize default health metrics
    db_health = models.ProjectHealth(
        project_id=db_project.id,
        schedule_status="Good",
        budget_status="Good",
        risk_status="Good"
    )
    db.add(db_health)
    db.commit()
    db.refresh(db_project)
    
    return db_project

@app.get("/api/projects", response_model=list[schemas.Project])
def get_projects(db: Session = Depends(database.get_db)):
    """Get all projects with their health metrics"""
    projects = db.query(models.Project).all()
    return projects

@app.get("/api/my-projects", response_model=list[schemas.Project])
def get_my_projects(email: str, db: Session = Depends(database.get_db)):
    """Get projects assigned to a specific user email"""
    # Note: In production we'd get email from JWT sub
    projects = db.query(models.Project).filter(models.Project.assigned_to_email == email).all()
    return projects

@app.put("/api/projects/{project_id}", response_model=schemas.Project)
def update_project(
    project_id: int,
    project: schemas.ProjectUpdate,
    db: Session = Depends(database.get_db)
):
    """Update an existing project partially"""
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Update project fields only if provided
    update_data = project.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_project, key, value)
    
    db.commit()
    db.refresh(db_project)
    return db_project

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, db: Session = Depends(database.get_db)):
    """Delete a project and its associated health metrics"""
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Delete associated health metrics first (cascade)
    db.query(models.ProjectHealth).filter(models.ProjectHealth.project_id == project_id).delete()
    
    # Delete associated payments, matters, tasks manually to ensure no FK errors
    db.query(models.PaymentSchedule).filter(models.PaymentSchedule.project_id == project_id).delete()
    db.query(models.MattersArising).filter(models.MattersArising.project_id == project_id).delete()
    db.query(models.ProjectTask).filter(models.ProjectTask.project_id == project_id).delete()
    
    # Delete the project
    db.delete(db_project)
    db.commit()
    
    return {"message": "Project deleted successfully", "id": project_id}

# Project Details Endpoints
@app.get("/api/projects/{project_id}/details", response_model=schemas.ProjectDetails)
def get_project_details(project_id: int, db: Session = Depends(database.get_db)):
    """Get complete project details including payments, matters, and tasks"""
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    return db_project

# Payment Schedule Endpoints
@app.post("/api/projects/{project_id}/payment", response_model=schemas.PaymentSchedule)
def create_payment(
    project_id: int,
    payment: schemas.PaymentScheduleCreate,
    db: Session = Depends(database.get_db)
):
    """Create a new payment milestone for a project"""
    # Verify project exists
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Create payment with all contract ledger fields
    db_payment = models.PaymentSchedule(
        project_id=project_id,
        category=payment.category,
        deliverable=payment.deliverable,
        phase=payment.phase,
        plan_date=payment.plan_date,
        planned_amount=payment.planned_amount,
        paid_amount=payment.paid_amount,
        status=payment.status,
        remark=payment.remark,
        payment_date=payment.payment_date,
        po_number=payment.po_number,
        invoice_number=payment.invoice_number
    )
    db.add(db_payment)
    db.commit()
    db.refresh(db_payment)
    
    return db_payment

@app.put("/api/payments/{payment_id}", response_model=schemas.PaymentSchedule)
def update_payment(
    payment_id: int,
    payment: schemas.PaymentScheduleUpdate,
    db: Session = Depends(database.get_db)
):
    """Update an existing payment record"""
    db_payment = db.query(models.PaymentSchedule).filter(models.PaymentSchedule.id == payment_id).first()
    if not db_payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Update only provided fields
    update_data = payment.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_payment, field, value)
    
    db.commit()
    db.refresh(db_payment)
    
    return db_payment

@app.delete("/api/projects/{project_id}/payments/{payment_id}")
def delete_payment(
    project_id: int,
    payment_id: int,
    db: Session = Depends(database.get_db)
):
    """Delete a payment record"""
    db_payment = db.query(models.PaymentSchedule).filter(
        models.PaymentSchedule.id == payment_id,
        models.PaymentSchedule.project_id == project_id
    ).first()
    
    if not db_payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    db.delete(db_payment)
    db.commit()
    
    return {"message": "Payment deleted successfully"}

# Matters Arising Endpoints
@app.post("/api/projects/{project_id}/matter", response_model=schemas.MattersArising)
def create_matter(
    project_id: int,
    matter: schemas.MattersArisingCreate,
    db: Session = Depends(database.get_db)
):
    """Create a new matter/issue for a project"""
    # Verify project exists
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Create matter
    # Create matter
    db_matter = models.MattersArising(
        project_id=project_id,
        issue_description=matter.issue_description,
        level=matter.level,
        action_updates=matter.action_updates,
        pic=matter.pic,
        target_date=matter.target_date,
        status=matter.status,
        date_closed=matter.date_closed,
        date_raised=matter.date_raised,
        remarks=matter.remarks
    )
    db.add(db_matter)
    db.commit()
    db.refresh(db_matter)
    
    return db_matter

@app.put("/api/matters/{matter_id}", response_model=schemas.MattersArising)
def update_matter(
    matter_id: int,
    matter: schemas.MattersArisingUpdate,
    db: Session = Depends(database.get_db)
):
    """Update an existing matter/issue"""
    db_matter = db.query(models.MattersArising).filter(models.MattersArising.id == matter_id).first()
    if not db_matter:
        raise HTTPException(status_code=404, detail="Matter not found")
    
    # Update only provided fields
    update_data = matter.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_matter, field, value)
    
    db.commit()
    db.refresh(db_matter)
    
    return db_matter

# Project Task Endpoints

@app.post("/api/projects/{project_id}/task", response_model=schemas.ProjectTask)
def create_task(
    project_id: int,
    task: schemas.ProjectTaskCreate,
    db: Session = Depends(database.get_db)
):
    """Create a new task for a project"""
    # Verify project exists
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Create task
    db_task = models.ProjectTask(
        project_id=project_id,
        task_name=task.task_name,
        start_date=task.start_date,
        end_date=task.end_date,
        duration=task.duration,
        completion_percentage=task.completion_percentage,
        completion_date=task.completion_date,
        status=task.status,
        parent_id=task.parent_id,
        assigned_to=task.assigned_to,
        priority=task.priority,
        description=task.description,
        tags=task.tags,
        order_index=task.order_index
    )
    db.add(db_task)
    db.commit()
    db.refresh(db_task)
    
    # Auto-recalculate project progress
    recalculate_project_progress(project_id, db)
    
    return db_task

@app.put("/api/projects/{project_id}/task/{task_id}", response_model=schemas.ProjectTask)
def update_task(
    project_id: int,
    task_id: int,
    task: schemas.ProjectTaskUpdate,
    db: Session = Depends(database.get_db)
):
    """Update a task"""
    db_task = db.query(models.ProjectTask).filter(
        models.ProjectTask.id == task_id,
        models.ProjectTask.project_id == project_id
    ).first()
    
    if not db_task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Update task fields
    update_data = task.dict(exclude_unset=True)
    
    # Auto-set completion date logic
    if "status" in update_data:
        if update_data["status"] == "Completed":
            # If completion_date is provided and not null, trust it.
            # If not provided, and not in DB, default to now.
            if "completion_date" in update_data and update_data["completion_date"]:
                pass # Use provided date
            elif not db_task.completion_date:
                update_data["completion_date"] = datetime.now().date()
        elif update_data["status"] != "Completed":
            update_data["completion_date"] = None

    for field, value in update_data.items():
        setattr(db_task, field, value)
    
    db.commit()
    db.refresh(db_task)
    
    # Auto-recalculate project progress
    recalculate_project_progress(project_id, db)
    
    return db_task

@app.delete("/api/projects/{project_id}/task/{task_id}")
def delete_task(
    project_id: int,
    task_id: int,
    db: Session = Depends(database.get_db)
):
    """Delete a task"""
    db_task = db.query(models.ProjectTask).filter(
        models.ProjectTask.id == task_id,
        models.ProjectTask.project_id == project_id
    ).first()
    
    if not db_task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    db.delete(db_task)
    db.commit()
    
    # Auto-recalculate project progress
    recalculate_project_progress(project_id, db)
    
    return {"message": "Task deleted successfully"}

# --- Task Comments Endpoints ---

@app.post("/api/tasks/{task_id}/comments", response_model=schemas.TaskComment)
def create_task_comment(
    task_id: int,
    comment: schemas.TaskCommentCreate,
    db: Session = Depends(database.get_db)
):
    """Add a comment to a task"""
    # Verify task exists
    db_task = db.query(models.ProjectTask).filter(models.ProjectTask.id == task_id).first()
    if not db_task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    db_comment = models.TaskComment(
        task_id=task_id,
        user_name=comment.user_name,
        content=comment.content,
        created_at=datetime.now().isoformat()
    )
    db.add(db_comment)
    db.commit()
    db.refresh(db_comment)
    return db_comment

@app.get("/api/projects/{project_id}/tasks_nested", response_model=list[schemas.ProjectTask])
def get_nested_tasks(
    project_id: int,
    db: Session = Depends(database.get_db)
):
    """Get tasks with subtasks for a project"""
    # Only fetch root tasks (where parent_id is None)
    tasks = db.query(models.ProjectTask).filter(
        models.ProjectTask.project_id == project_id,
        models.ProjectTask.parent_id == None
    ).all()
    return tasks

# --- Excel Import/Export Endpoints ---

@app.get("/api/payments/template")
def get_payment_template():
    # Create a workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Schedule Template"
    
    # Define headers
    headers = ["Deliverable", "Phase", "Plan Date (YYYY-MM-DD)", "Planned Amount", "Remarks"]
    ws.append(headers)
    
    # Add sample row
    ws.append(["Milestone 1", "Phase 1", "2026-01-31", 5000.00, "Initial payment"])
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
        
    return FileResponse(tmp_path, filename="payment_schedule_template.xlsx", media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.post("/api/projects/{project_id}/payments/import")
async def import_payments(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db)
):
    # Verify project exists
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # verify file extension
    if not file.filename.endswith('.xlsx'):
         raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file (.xlsx)")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        imported_count = 0
        
        # Iterate rows, skipping header (min_row=2)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: # Skip empty rows
                continue
                
            # Map columns (Deliverable, Phase, Date, Amount, Remarks)
            deliverable = row[0]
            phase = row[1]
            plan_date_raw = row[2]
            planned_amount = row[3]
            remark = row[4]
            
            # Date Parsing
            plan_date = None
            if plan_date_raw:
                if isinstance(plan_date_raw, datetime) or isinstance(plan_date_raw, date):
                     plan_date = plan_date_raw
                     if isinstance(plan_date, datetime):
                         plan_date = plan_date.date()
                elif isinstance(plan_date_raw, str):
                    try:
                        plan_date = datetime.strptime(plan_date_raw, "%Y-%m-%d").date()
                    except ValueError:
                        pass # Valid validation in real app needed
            
            # Create payment record
            db_payment = models.PaymentSchedule(
                project_id=project_id,
                deliverable=str(deliverable),
                phase=str(phase) if phase else None,
                plan_date=plan_date,
                planned_amount=float(planned_amount) if planned_amount else 0.0,
                status="Not Paid",
                remark=str(remark) if remark else None
            )
            db.add(db_payment)
            imported_count += 1
            
        db.commit()
        return {"message": f"Successfully imported {imported_count} payment records"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")
